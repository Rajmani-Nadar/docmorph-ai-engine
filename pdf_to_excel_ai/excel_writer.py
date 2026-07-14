from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from .utils import EXPECTED_COLUMNS
except ImportError:  # pragma: no cover - fallback for direct script execution
    from utils import EXPECTED_COLUMNS


def write_excel(records: list[dict[str, str]], output_path: Path, logger: logging.Logger) -> Path:
    """Write extracted records to an Excel workbook with formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Determine headers dynamically from the first record when available
    if records:
        # Preserve column order from the record keys
        first = records[0]
        if isinstance(first, dict):
            columns = list(first.keys())
        else:
            columns = EXPECTED_COLUMNS
    else:
        columns = EXPECTED_COLUMNS
    frame = pd.DataFrame(records, columns=columns)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Student Records"

    for column_index, column_name in enumerate(columns, start=1):
        sheet.cell(row=1, column=column_index, value=column_name)

    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value="" if value is None else str(value))

    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPECTED_COLUMNS))}{sheet.max_row}"

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, _ in enumerate(columns, start=1):
        column_cells = [sheet.cell(row=row_num, column=column_index).value for row_num in range(1, sheet.max_row + 1)]
        max_length = max(len(str(cell or "")) for cell in column_cells)
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width

    workbook.save(output_path)
    logger.info("Excel Updated")
    return output_path
